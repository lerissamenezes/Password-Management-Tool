from tkinter import *
from tkinter import messagebox
import random
import pyperclip
from datetime import datetime
import bcrypt  
import sys
import openpyxl
from openpyxl.workbook import Workbook as wb
import rsa
import psutil
import subprocess
import os
import string
import xlsxwriter
import xlrd

#-----------------------------VARIABLE/FUCNTION INITIALISATION----------------------------#
NAVY = '#0A043C'
GREY = '#bbbbbb'
BEIGE = '#ffe3d8'
current_directory = os.getcwd()
loc = ("D:/MINI/pasheet.xlsx")
#---------------------------FOR INITIALIZING KEY------------------------------------------#
alpha='abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-=!@#$%^&*()_+=' #crete key list
l=list(alpha)
random.shuffle(l)
keyx="".join(l)
#---------------------------ENCRYPTION-------------------------------------------------#
def encryptorx(message,key):
  cipher=""
  for i in message:
    if i==" ":
      cipher+=i
    else:
      cipher+=key[alpha.index(i)]
  print("Encrypted password:",cipher)
  password_entry.delete(0, END)
  password_entry.insert(0, cipher)
#---------------------------DECRYPTION-------------------------------------------------#
def decrypt(cipher,key):
    plaintext=""
    for i in cipher:
        if i==" ":
            plaintext+=i
        else:
            plaintext+=alpha[key.index(i)]
    print("Decrypted Password:",plaintext)
    pyperclip.copy(plaintext)
    messagebox.showinfo("SUCCESS","Password decrypted & copied to clipboard!")
    show_entry1.delete(0, END)
    show_entry1.insert(0, plaintext)
#---------------------------OPEN_XL STUFF-------------------------------------------------#
def open_xl():
    doc = subprocess.Popen(["start", "/WAIT", "pasheet.xlsx"], shell=True)   
    doc.poll()                                                           
    #psutil.Process(doc.pid).get_children()[0].kill()                     
    doc.poll()                                                           
def close_win(top):
   top.destroy()
def insert_val(e):
   if e.get()== 'password':
       open_xl()
   else:
       messagebox.showinfo("ERROR","Wrong password")
def insert_val1(e):
   if e.get()== 'password':
    wrkbk = openpyxl.load_workbook("pasheet.xlsx")
    web= show_entry.get()
    sh = wrkbk.active
  
    # iterate through excel and display data
    for i in range(1, sh.max_row+1):
       for j in range(1, sh.max_column+1):
           cell_obj = sh.cell(row=i, column=1)
           cell_pass= sh.cell(row=i, column=3)
           if cell_obj.value == web:
                pyperclip.copy(cell_pass.value)
                messagebox.showinfo("SUCCESS","Copied pass to clipboard")
                break
   else:
       messagebox.showinfo("ERROR","Wrong password")   
 #def insert_show(w):
  #  for w.get in range:       
       
#----------------------------ENCRYPTION -----------------------------------#

#---------------------------WORKBOOK---------------------------------------------#

workbook = openpyxl.load_workbook(current_directory+'\pasheet.xlsx')
worksheet = workbook.active

# ---------------------------- PASSWORD GENERATOR ------------------------------- #
def random_password():

    letters = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    numbers = list('0123456789')
    symbols = list('!@#$%^&*()_+')
    letters_lower = list(map(str.lower,letters))
    letters.extend(letters_lower)

    #Return a number between a and b (both included):
    num_letters = random.randint(8,10)
    num_numbers = random.randint(1,2)
    num_symbols = random.randint(1,2)

    # Creating password and concat to a list
    rand_letters = [random.choice(letters) for i in range(num_letters)]
    rand_numbers = [random.choice(numbers) for i in range(num_numbers)]
    rand_symbols = [random.choice(symbols) for i in range(num_symbols)]

    created_password = rand_letters + rand_numbers + rand_symbols
    password_entry.delete(0, END)
    password_entry.insert(0, created_password)
     
   
    random.shuffle(created_password)
    created_password = ''.join(created_password)
    print("Random Password generated:",created_password)
    encryptorx(created_password,keyx)

    #publicKey, privateKey = rsa.newkeys(512)

    #encrypted_password = rsa.encrypt(created_password.encode(),publicKey)
    
    #password_entry.insert(0, created_password)#3Jwn(URRK)aE0h
    

# ---------------------------- SAVE PASSWORD ------------------------------- #
def saved_entries():

    user_website = website_entry.get()
    user_email = email_entry.get()
    user_password = password_entry.get()

    #publicKey, privateKey = rsa.newkeys(512)

   # encrypted_password = rsa.encrypt(user_password.encode(),publicKey)

    #ENCRYPTING FROM ABOVE CLASS
    #enc=Encryptor(user_password)
    #enc.encrypt_string()
    #-----WRITING INTO EXCEL SHEET---------#
    worksheet.append([user_website,user_email,user_password])  
    workbook.save(current_directory+'\pasheet.xlsx')
    website_entry.delete(0,END)
    email_entry.delete(0,END)
    password_entry.delete(0,END)
    
#-----------------------------RETRIEVE ALL PASS FUNCTION--------------------#
def popupwin():
    top= Toplevel(root)
    top.geometry("500x250")

    top.config(padx=50, pady=50, bg=NAVY)
    entry= Entry(top, width= 25,show='●')
    entry.pack()

    Button(top,text= "Sign in", command= lambda:insert_val(entry)).pack(pady= 5,side=TOP)
   
    button= Button(top, text="Cancel", command=lambda:close_win(top))
    button.pack(pady=5, side= TOP)


#-----------------------------RETRIEVE PASS FUNCTION--------------------#
def show():
    top= Toplevel(root)
    top.geometry("500x250")

    top.config(padx=50, pady=50, bg=NAVY)
    entry= Entry(top, width= 25,show='●')
    entry.pack()

    Button(top,text= "Sign in", command= lambda:insert_val(entry)).pack(pady= 5,side=TOP)
   
    button= Button(top, text="Cancel", command=lambda:close_win(top))
    button.pack(pady=5, side= TOP)

def show1():         
    top= Toplevel(root)
    top.geometry("500x250")

    top.config(padx=50, pady=50, bg=NAVY)
    entry= Entry(top, width= 25,show='●')
    entry.pack()

    Button(top,text= "Sign in", command= lambda:insert_val1(entry)).pack(pady= 5,side=TOP)
     
    button= Button(top, text="Cancel", command=lambda:close_win(top))
    button.pack(pady=5, side= TOP)
    
            
            

# ---------------------------- UI SETUP ------------------------------- #

root = Tk()
root.title("Password Manager")
root.config(padx=50, pady=50, bg=NAVY)

canvas = Canvas(height=200, width=200, bg=NAVY, highlightthickness=0)
img = PhotoImage(file='logo2.png')
canvas.create_image(100, 100, image=img)
canvas.grid(row=0,column=1)

# ROW 1
website_label = Label(text='Website:', bg=NAVY, fg=BEIGE)
website_label.grid(row=1,column=0,sticky="W")

website_entry = Entry(font=('Arial',15))
website_entry.grid(row=1,column=1, columnspan=2,sticky="EW")
website_entry.focus()


# ROW 2
email_label = Label(text='Email/Username:', bg=NAVY, fg=BEIGE)
email_label.grid(row=2,column=0,sticky="W")

email_entry = Entry(font=('Arial',15))
email_entry.grid(row=2,column=1, columnspan=2,sticky="EW")
email_entry.insert(0, '')

# ROW 3
password_label = Label(text='Password:', bg=NAVY, fg=BEIGE)
password_label.grid(row=3,column=0,sticky="W")

password_entry = Entry(font=('Arial',15),show='●')
password_entry.grid(row=3,column=1,sticky="EW")

password_button = Button(text='Generate Password', bg=GREY, command= random_password)
password_button.grid(row=3,column=2,sticky="EW")

# ROW 4
button = Button(text='Add', bg=GREY, command=saved_entries)
button.grid(row=6,column=1,columnspan=2,sticky="EW")
button.config(pady=2)

#ROW 5
button = Button(text='Retrieve Passwords', bg=GREY,command=popupwin)
button.grid(row=7,column=1,columnspan=2,sticky="EW")
button.config(pady=2)

#ROW 6
show_label = Label(text='Which website do you want to log into? ', bg=NAVY, fg=BEIGE)
show_label.grid(row=4,column=0,sticky="W")

show_entry= Entry(font=('Arial',15))
show_entry.grid(row=4,column=1,sticky="EW")

show_button = Button(text='Show Password', bg=GREY, command=show1)
show_button.grid(row=4,column=2,sticky="EW")

show_label = Label(text='Decrypt ', bg=NAVY, fg=BEIGE)
show_label.grid(row=5,column=0,sticky="W")

show_entry1= Entry(font=('Arial',15))
show_entry1.grid(row=5,column=1,sticky="EW")
show_button1 = Button(text='Decrypt', bg=GREY, command=lambda:decrypt(show_entry1.get(),keyx))
show_button1.grid(row=5,column=2,sticky="EW")

root.mainloop()

